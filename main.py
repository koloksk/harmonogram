import json
import re
import sys
from datetime import datetime, timedelta
from typing import List, Tuple, Optional, Dict, Any


def strip_or_empty(s: Optional[str]) -> str:
    return (s or "").strip()

# -----------------------------
# Daty i miesiące (z wariantami mojibake)
# -----------------------------
MONTH_VARIANTS: Dict[str, List[str]] = {
    "01": ["stycznia"],
    "02": ["lutego"],
    "03": ["marca"],
    "04": ["kwietnia"],
    "05": ["maja"],
    "06": ["czerwca"],
    "07": ["lipca"],
    "08": ["sierpnia"],
    "09": ["wrzesnia", "września", "wrze?nia"],
    "10": ["października", "pazdziernika", "pa?dziernika"],
    "11": ["listopada"],
    "12": ["grudnia"],
}


def parse_polish_date_cell(cell: str) -> Optional[str]:
    if not cell:
        return None
    s = cell.lower()
    year_match = re.search(r"(20\d{2})", s)
    if not year_match:
        return None
    year = int(year_match.group(1))
    day_match = re.search(r"\b(\d{1,2})\b", s)
    if not day_match:
        return None
    day = int(day_match.group(1))
    month_num = None
    for m_num, variants in MONTH_VARIANTS.items():
        for v in variants:
            if v in s:
                month_num = m_num
                break
        if month_num:
            break
    if not month_num:
        parts = re.findall(r"[a-zA-ZąćęłńóśźżĄĆĘŁŃÓŚŹŻ\?]+", s)
        if len(parts) >= 2:
            mid = parts[1].replace("?", "z")
            for m_num, variants in MONTH_VARIANTS.items():
                if any(mid == v for v in variants):
                    month_num = m_num
                    break
    if not month_num:
        return None
    try:
        dt = datetime(year, int(month_num), day)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None


def fmt_hhmm(h: int, m: int) -> str:
    return f"{h:02d}:{m:02d}"


def add_minutes(hhmm: str, minutes: int) -> str:
    h, m = map(int, hhmm.split(":"))
    dt = datetime(2000, 1, 1, h, m) + timedelta(minutes=minutes)
    return dt.strftime("%H:%M")


# -----------------------------
# Parsowanie treści komórki: typ, prowadzący, sala (bez zdalności)
# -----------------------------
TYPE_KEYWORDS = [
    ("LAB", "LAB"),
    (" PROJEKT", "PROJEKT"),
    ("PROJEKT", "PROJEKT"),
    (" W", "W"),
    (" ?W", "ĆW"),
    (" ?w", "ĆW"),
    (" K", "K"),
    ("wykład", "W"),
    ("konwersatorium", "K"),
    ("ćw", "ĆW"),
]

LECTURER_PATTERN = re.compile(r"\b(?:dr hab\.|dr|mgr|prof\.?)(?: [a-zA-ZąćęłńóśźżĄĆĘŁŃÓŚŹŻ\.]+){1,3}")
ROOM_PATTERNS = [
    re.compile(r"\b\d+\.\d+\s*(?:CP|CI|CsH)\b", re.IGNORECASE),
    re.compile(r"\b(?:CP|CI|CsH)\s*\d+\b", re.IGNORECASE),
    re.compile(r"\baula[^;\n]*", re.IGNORECASE),
    re.compile(r"\bsala[^;\n]*", re.IGNORECASE),
]


def parse_cell_details(text: str) -> Dict[str, Any]:
    s = strip_or_empty(text)
    base = {"title": None, "type": None, "lecturers": [], "location": None, "isRemote": False, "color": None, "raw": s}
    if not s:
        return base
    s_up = " " + s.upper()
    for key, val in TYPE_KEYWORDS:
        if key.upper() in s_up:
            base["type"] = val
            break
    lecturers = LECTURER_PATTERN.findall(s)
    if lecturers:
        base["lecturers"] = lecturers
    for pat in ROOM_PATTERNS:
        m = pat.search(s)
        if m:
            base["location"] = m.group(0).strip()
            break
    # Nie ustawiamy tutaj isRemote/color – decyduje kolor czcionki w XLSX
    t = s
    t = re.sub(r"\b\d+\s*/\s*\d+\b", " ", t)
    for lec in lecturers:
        t = t.replace(lec, " ")
    if base["location"]:
        t = t.replace(base["location"], " ")
    for key, _ in TYPE_KEYWORDS:
        t = t.replace(key.strip(), " ")
    t = re.sub(r"\s+", " ", t).strip(" ;,\n\t")
    if t:
        base["title"] = t
    return base


# -----------------------------
# XLSX: odczyt całego harmonogramu (wszystkie kierunki)
# -----------------------------

def parse_xlsx_to_events(xlsx_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as ie:
        raise RuntimeError("Brak biblioteki 'openpyxl'. Zainstaluj: pip install openpyxl") from ie

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    max_col = ws.max_column
    max_row = ws.max_row

    # Forward-fill nagłówki w wierszach 1..4
    headers: Dict[int, Dict[str, Optional[str]]] = {}
    last_zjazd = last_date_raw = last_day = last_program = None

    def to_date_iso(val: Optional[object]) -> Optional[str]:
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        try:
            from datetime import date as _dt_date
            if isinstance(val, _dt_date):
                return val.strftime("%Y-%m-%d")
        except Exception:
            pass
        return parse_polish_date_cell(str(val))

    for c in range(1, max_col + 1):
        zjazd = ws.cell(row=1, column=c).value
        date_v = ws.cell(row=2, column=c).value
        day_v = ws.cell(row=3, column=c).value
        program_v = ws.cell(row=4, column=c).value
        if zjazd is not None and str(zjazd).strip():
            last_zjazd = str(zjazd).strip()
        if date_v is not None and str(date_v).strip():
            last_date_raw = date_v
        if day_v is not None and str(day_v).strip():
            last_day = str(day_v).strip()
        if program_v is not None and str(program_v).strip():
            last_program = str(program_v).strip()
        headers[c] = {
            "zjazd": last_zjazd,
            "date": to_date_iso(last_date_raw) if last_date_raw is not None else None,
            "day": last_day,
            "program": last_program,
        }

    # Mapa scalen
    merged_map: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
    for rng in ws.merged_cells.ranges:
        min_r, min_c, max_r, max_c = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        for rr in range(min_r, max_r + 1):
            for cc in range(min_c, max_c + 1):
                merged_map[(rr, cc)] = (min_r, min_c, max_r, max_c)

    def time_from_row(rr: int) -> Optional[str]:
        # Wiersze 6..57: sloty 15-min od 08:00
        if rr < 6:
            return None
        idx = rr - 6
        h = 8 + (idx * 15) // 60
        m = (idx * 15) % 60
        return f"{h:02d}:{m:02d}"

    def is_cell_text_red(cell) -> bool:
        # tylko kolor czcionki decyduje o zdalności
        col = cell.font.color
        if col is None:
            return False
        try:
            if getattr(col, 'type', None) == 'rgb' and getattr(col, 'rgb', None):
                val = col.rgb.upper()
                return val.endswith('FF0000') or val in ('FFFF0000', 'FF0000', '00FF0000')
            # indeks 3 to często czerwony w palecie xls
            if getattr(col, 'indexed', None) in (3,):
                return True
        except Exception:
            return False
        return False

    events: List[Dict[str, Any]] = []

    for c in range(1, max_col + 1):
        meta = headers.get(c) or {}
        date_iso = meta.get("date")
        day_name = meta.get("day")
        program = meta.get("program")
        zjazd = meta.get("zjazd")
        if not date_iso or not program:
            continue
        r = 6
        while r <= min(max_row, 57):
            tl = merged_map.get((r, c))
            height = 1
            top_left_r, top_left_c = r, c
            if tl:
                min_r, min_c, max_r, max_c = tl
                if (r, c) != (min_r, min_c):
                    r += 1
                    continue
                height = max_r - min_r + 1
                top_left_r, top_left_c = min_r, min_c
            cell_obj = ws.cell(row=top_left_r, column=top_left_c)
            cell_val = cell_obj.value
            text = strip_or_empty(str(cell_val) if cell_val is not None else "")
            if text:
                details = parse_cell_details(text)
                start_time = time_from_row(r)
                end_time = add_minutes(start_time, height * 15) if start_time else None
                remote = is_cell_text_red(cell_obj)
                events.append({
                    "zjazd": zjazd,
                    "program": program,
                    "date": date_iso,
                    "day": day_name,
                    "startTime": start_time,
                    "endTime": end_time,
                    "title": details["title"],
                    "type": details["type"],
                    "lecturers": details["lecturers"],
                    "location": details["location"],
                    "isRemote": remote,
                    "color": ("red" if remote else None),
                    "raw": details["raw"],
                })
                r += height
            else:
                r += 1

    events.sort(key=lambda e: (e.get("date") or "", e.get("program") or "", e.get("startTime") or ""))
    return events


# -----------------------------
# Zapis JSON (tylko XLSX)
# -----------------------------

def export_schedule_to_json(input_path: str, out_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    lower = input_path.lower()
    if not lower.endswith((".xlsx", ".xlsm")):
        raise ValueError("Obsługiwane są tylko pliki XLSX/XLSM.")
    events = parse_xlsx_to_events(input_path, sheet_name)
    payload = {
        "source_file": input_path,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "program": "ALL",
        "events": events,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return {"events_count": len(events), "out_file": out_path}


# -----------------------------
# Uruchomienie z CLI
# -----------------------------
if __name__ == "__main__":
    args = sys.argv[1:]
    if not args:
        print("Użycie: python main.py <plik.xlsx|plik.xlsm> [nazwa_arkusza]")
        sys.exit(1)
    input_path = args[0]
    sheet = args[1] if len(args) >= 2 else None
    out_path = "harmonogram.json"
    try:
        result = export_schedule_to_json(input_path, out_path, sheet_name=sheet)
        print(f"Zapisano {result['events_count']} pozycji do {result['out_file']}")
    except Exception as e:
        print(f"Błąd: {e}")
        sys.exit(1)
